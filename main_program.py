import sys
import re
import numpy as np
import pandas as pd
from ast import literal_eval
from pathlib import Path
from openpyxl import load_workbook

from nltk.tokenize import word_tokenize

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import QtCore, QtGui, QtWidgets

##  https://wikidocs.net/35482
##  https://yonoo88.tistory.com/1209
##  http://pythonstudy.xyz/python/article/108-PyQt-QtDesigner
##  https://wikidocs.net/21952

#GUI = uic.loadUiType("GUI_ver3.ui")[0]

class NoVendor(Exception):
    def __init__(self):
        super().__init__('Please Select Vendor')

class Unmatch(Exception):
    def __init__(self):
        super().__init__('Please Check Format')

class NoFile(Exception):
    def __init__(self):
        super().__init__('No File Exist')



class Thread(QThread):
    threadEvent = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__()
        self.n = 0
        self.main = parent
        self.isRun = False
        self.functions = {0:'.',
                          1:'..',
                          2:'...'}
    def run(self):
        while self.isRun:
            a = str('Processing'+self.functions[self.n % 3])
            self.threadEvent.emit(a)
            self.n += 1
            self.msleep(1000)


class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(760, 450)
        self.inputpath = QtWidgets.QLineEdit(Dialog)
        self.inputpath.setEnabled(True)
        self.inputpath.setGeometry(QtCore.QRect(30, 130, 291, 20))
        self.inputpath.setAcceptDrops(False)
        self.inputpath.setReadOnly(True)
        self.inputpath.setObjectName("inputpath")
        self.label = QtWidgets.QLabel(Dialog)
        self.label.setGeometry(QtCore.QRect(30, 110, 91, 16))
        self.label.setObjectName("label")
        self.selectinput = QtWidgets.QPushButton(Dialog)
        self.selectinput.setGeometry(QtCore.QRect(250, 100, 75, 23))
        self.selectinput.setObjectName("selectinput")
        self.algorithmstart = QtWidgets.QPushButton(Dialog)
        self.algorithmstart.setGeometry(QtCore.QRect(610, 360, 91, 23))
        self.algorithmstart.setObjectName("algorithmstart")
        self.algorithmsave = QtWidgets.QPushButton(Dialog)
        self.algorithmsave.setGeometry(QtCore.QRect(610, 400, 91, 23))
        self.algorithmsave.setObjectName("algorithmsave")
        self.showtable = QtWidgets.QTableWidget(Dialog)
        self.showtable.setGeometry(QtCore.QRect(20, 180, 561, 251))
        self.showtable.setDragDropOverwriteMode(False)
        self.showtable.setAlternatingRowColors(True)
        self.showtable.setRowCount(0)
        self.showtable.setObjectName("showtable")
        self.showtable.setColumnCount(4)
        item = QtWidgets.QTableWidgetItem()
        self.showtable.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.showtable.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.showtable.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.showtable.setHorizontalHeaderItem(3, item)
        self.label_2 = QtWidgets.QLabel(Dialog)
        self.label_2.setEnabled(True)
        self.label_2.setGeometry(QtCore.QRect(450, 30, 151, 16))
        self.label_2.setTextFormat(QtCore.Qt.PlainText)
        self.label_2.setObjectName("label_2")
        self.groupBox = QtWidgets.QGroupBox(Dialog)
        self.groupBox.setGeometry(QtCore.QRect(450, 80, 131, 80))
        self.groupBox.setObjectName("groupBox")
        self.vendorGE = QtWidgets.QRadioButton(self.groupBox)
        self.vendorGE.setGeometry(QtCore.QRect(10, 20, 101, 16))
        self.vendorGE.setObjectName("vendorGE")
        self.vendorPhilips = QtWidgets.QRadioButton(self.groupBox)
        self.vendorPhilips.setGeometry(QtCore.QRect(10, 40, 101, 16))
        self.vendorPhilips.setObjectName("vendorPhilips")
        self.vendorNihon = QtWidgets.QRadioButton(self.groupBox)
        self.vendorNihon.setGeometry(QtCore.QRect(10, 60, 111, 16))
        self.vendorNihon.setObjectName("vendorNihon")
        self.selectadditional = QtWidgets.QPushButton(Dialog)
        self.selectadditional.setGeometry(QtCore.QRect(660, 20, 75, 23))
        self.selectadditional.setObjectName("selectadditional")
        self.additionalpath = QtWidgets.QLineEdit(Dialog)
        self.additionalpath.setGeometry(QtCore.QRect(450, 50, 281, 20))
        self.additionalpath.setReadOnly(True)
        self.additionalpath.setObjectName("additionalpath")
        self.additionalstart = QtWidgets.QPushButton(Dialog)
        self.additionalstart.setGeometry(QtCore.QRect(590, 140, 75, 23))
        self.additionalstart.setObjectName("additionalstart")
        self.additionaltext = QtWidgets.QLabel(Dialog)
        self.additionaltext.setGeometry(QtCore.QRect(590, 120, 161, 16))
        self.additionaltext.setText("")
        self.additionaltext.setObjectName("additionaltext")
        self.algorithmtext = QtWidgets.QLabel(Dialog)
        self.algorithmtext.setGeometry(QtCore.QRect(590, 340, 161, 20))
        self.algorithmtext.setText("")
        self.algorithmtext.setObjectName("algorithmtext")
        self.Statement = QtWidgets.QLabel(Dialog)
        self.Statement.setGeometry(QtCore.QRect(30, 20, 151, 21))
        self.Statement.setObjectName("Statement")
        self.progressBar = QtWidgets.QProgressBar(Dialog)
        self.progressBar.setGeometry(QtCore.QRect(590, 310, 118, 23))
        self.progressBar.setProperty("value", 0)
        self.progressBar.setTextVisible(True)
        self.progressBar.setObjectName("progressBar")

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Mapping of EKG Statement to CDM Concept Code"))
        self.label.setText(_translate("Dialog", "Input File Path"))
        self.selectinput.setText(_translate("Dialog", "Select"))
        self.algorithmstart.setText(_translate("Dialog", "Run"))
        self.algorithmsave.setText(_translate("Dialog", "Save Result"))
        item = self.showtable.horizontalHeaderItem(0)
        item.setText(_translate("Dialog", "Vendor"))
        item = self.showtable.horizontalHeaderItem(1)
        item.setText(_translate("Dialog", "Statement"))
        item = self.showtable.horizontalHeaderItem(2)
        item.setText(_translate("Dialog", "concept ID"))
        item = self.showtable.horizontalHeaderItem(3)
        item.setText(_translate("Dialog", "concept name"))
        self.label_2.setText(_translate("Dialog", "Additional Dictionary Path"))
        self.groupBox.setTitle(_translate("Dialog", "Vendor"))
        self.vendorGE.setText(_translate("Dialog", "GE"))
        self.vendorPhilips.setText(_translate("Dialog", "Philips"))
        self.vendorNihon.setText(_translate("Dialog", "Nihon Kohden"))
        self.selectadditional.setText(_translate("Dialog", "Select"))
        self.additionalstart.setText(_translate("Dialog", "Add"))
        self.Statement.setText(_translate("Dialog", "Initializing..."))



#GUI
class WindowClass(QDialog, Ui_Dialog) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)

        self.show()
        self.th = Thread(parent=self)
        self.th.threadEvent.connect(self.threadEventHandler)

        # Event Handling
        self.selectinput.clicked.connect(self.selectinput_handle)
        self.selectadditional.clicked.connect(self.selectadditional_handle)
        self.additionalstart.clicked.connect(self.additionalstart_handle)
        self.algorithmstart.clicked.connect(self.algorithm_handle)
        self.algorithmsave.clicked.connect(self.algorithmsave_handle)

        self.threadStart

        self.GE_rule = EKG_rule()
        self.GE_rule.Data_Load('dictionary_data/GE_mapping.xlsx')
        self.Philips_rule = EKG_rule()
        self.Philips_rule.Data_Load('dictionary_data/Philips_mapping.xlsx')
        self.Nihon_rule = EKG_rule()
        self.Nihon_rule.Data_Load('dictionary_data/Kohden_mapping.xlsx')

        self.inputpathtext = None
        self.additionalpathtext = None
        self.output_data = None

        self.threadStop

        self.Statement.setText('')
        self.Statement.repaint()

    @pyqtSlot(str)
    def threadEventHandler(self, n):
        self.Statement.setText(str(n))
        self.Statement.repaint()

    @pyqtSlot()
    def threadStart(self):
        self.th.isRun = True
        self.th.run()

    @pyqtSlot()
    def threadStop(self):
        self.th.isRun = False

    def selectinput_handle(self): # select input path
        fname = QFileDialog.getOpenFileName(self, 'Select Input File', './', 'Excel Worksheet File(*.xlsx);;CSV file(*.csv)')
        self.inputpathtext = fname[0]
        self.inputpath.setText(self.inputpathtext)

    def selectadditional_handle(self):  # select additional dictionary path
        fname = QFileDialog.getOpenFileName(self, 'Select Additional Dictionary File', './', 'Excel Worksheet File(*.xlsx)')
        self.additionalpathtext = fname[0]
        self.additionalpath.setText(self.additionalpathtext)

    def additionalstart_handle(self):  # adding additional dictionary (progress will show in self.additionaltext)
        self.additionaltext.setText('Processing...')
        self.additionaltext.repaint()

        try:
            if self.additionalpathtext == None: raise NoFile
            if self.vendorGE.isChecked():
                    try: self.GE_rule.additional_Data_Load(Path(self.additionalpathtext))
                    except:  raise Unmatch
            elif self.vendorPhilips.isChecked():
                    try: self.Philips_rule.additional_Data_Load(Path(self.additionalpathtext))
                    except: raise Unmatch
            elif self.vendorNihon.isChecked():
                    try: self.Nihon_rule.additional_Data_Load(Path(self.additionalpathtext))
                    except: raise Unmatch
            else: raise NoVendor

        except NoVendor:
            self.additionaltext.setText('Error: Please Select Vendor')
        except Unmatch:
            self.additionaltext.setText('Error: Please Check Format')
        except NoFile:
            self.additionaltext.setText('Error: No File Exist')

        else:
            self.additionaltext.setText('Done')
            self.additionalpathtext = None
            self.additionalpath.setText('')



    def algorithm_handle(self):  # Algorithm (print to self.showtable, progress will show in self.algorithmtext)
        self.algorithmtext.setText('Processing...')
        self.algorithmtext.repaint()
        self.progressBar.setValue(0)
        try:
            if self.inputpathtext == None: raise NoFile
            ext = QFileInfo(self.inputpathtext).suffix()
            if ext == 'csv':
                    try:
                        data = pd.read_csv(Path(self.inputpathtext))
                        input_vendor = data['Vendor']
                        input_source = data['Diagnosis Statement']
                    except:  raise Unmatch
            elif ext == 'xlsx':
                    try:
                        load_wb = load_workbook(Path(self.inputpathtext))
                        load_ws = load_wb.active  # first sheet
                        data = pd.DataFrame([[i.value for i in j] for j in load_ws.rows])
                        header = data.iloc[0]
                        data = data[1:]
                        data.rename(columns=header, inplace=True)
                        data.reset_index(drop=True, inplace=True)
                        input_vendor = data['Vendor']
                        input_source = data['Diagnosis Statement']
                    except: raise Unmatch
            else: raise NoVendor
            total = data.shape[0]
            
        except NoVendor:
            self.algorithmtext.setText('Error: Check File Extension')
        except Unmatch:
            self.algorithmtext.setText('Error: Unsuitable Format')
        except NoFile:
            self.algorithmtext.setText('Error: No File Exist')

        else:
            item = []
            self.progressBar.setMaximum(total)
            for i in range(0, total):
                self.progressBar.setValue(i)
                # bug fix, 2020.11.03 shc
                if input_vendor[i] == ('GE' or 'GE Healthcare'):
                    id, name = self.GE_rule.Check_if_any(literal_eval(input_source[i]))
                elif input_vendor[i] == 'Philips':
                    id, name = self.Philips_rule.Get_similar(literal_eval(input_source[i]))
                elif input_vendor[i] == ('Nihon Kohden' or 'Kohden' or 'Nihon'):
                    id, name = self.Nihon_rule.Get_similar(literal_eval(input_source[i]))
                else:
                    ## 201007 Should be filled.
                    ## Average 비교, 후 매핑
                    id = []
                    name = []
                id = [second for first in id for second in first]
                name = [second for first in name for second in first]
                item.append([input_vendor[i], input_source[i], list(set(id)), list(set(name))])

            self.progressBar.setValue(total)
            self.algorithmtext.setText('Printing Results...')
            self.algorithmtext.repaint()
            self.output_data = pd.DataFrame(data=item, columns=['Vendor', 'Statement', 'concept ID', 'concept name'])
            self.showtable.setRowCount(total)
            for row in range(self.output_data.shape[0]):
                for col in range(self.output_data.shape[1]):
                    self.showtable.setItem(row, col, QTableWidgetItem(str(item[row][col])))

            self.algorithmtext.setText('Done')



    def algorithmsave_handle(self):  # saving the result of algorithm
        fname = QFileDialog.getSaveFileName(self, 'Save Table Data', './', 'Excel Worksheet File(*.xlsx);;CSV file(*.csv)')
        ext = QFileInfo(fname[0]).suffix()
        if ext == 'csv':
            self.output_data.to_csv(fname[0], sep=',', index=False)
        elif ext == 'xlsx':
            self.output_data.to_excel(fname[0], index=False)
		



# EKG Algorithm
class EKG_rule:
    def __init__(self):
        self.Rule = None
        self.should_not_use = None
        self.comment3 = None
        self.if_any_match = None
        self.X = None
        self.y = None
        self.tfidf = None
        self.tfidf_matrix = None
        self.index_list = None
        self.comment4 = None

    def Data_Load(self, filename, data_only=True):
        load_wb = load_workbook(filename, data_only)
        rule = []
        load_ws = load_wb.active  # first sheet
        data = pd.DataFrame([[i.value for i in j] for j in load_ws.rows])
        header = data.iloc[0]
        data = data[1:]
        data.rename(columns=header, inplace=True)
        data.reset_index(drop=True, inplace=True)

        self.Rule = data[['source_name', 'condition_concept_id', 'concept_name']]

        comma = self.Rule['source_name'].str.replace(pat=',', repl='', regex=False)

        self.X = comma.str.replace(pat='-', repl=' ', regex=False)
        X2 = comma.str.replace(pat='-', repl='', regex=False)
        self.X = self.X.append(X2, ignore_index=False)
        self.X = self.X.str.lower()
        self.X.drop_duplicates(inplace=True)

        # self.X = self.X.str.replace(pat=',', repl='', regex=False)
        self.y = self.Rule[['condition_concept_id', 'concept_name']]

        self.index_list = self.X.index.tolist()

        # self.tfidf = TfidfVectorizer()
        self.tfidf = CountVectorizer(stop_words=['***', ','], tokenizer=word_tokenize)

        self.tfidf_matrix = self.tfidf.fit_transform(self.X)
        # print(self.tfidf_matrix.shape)

        self.should_not_use = list(comma[data['should_not_use'] == 2].str.lower())
        self.comment3 = list(comma[data['comment'] == 3].str.lower())
        self.comment4 = list(comma[data['comment'] == 4].str.lower())

    def additional_Data_Load(self, filename, data_only=True):
        load_wb = load_workbook(filename, data_only)
        rule = []
        load_ws = load_wb.active  # first sheet
        data = pd.DataFrame([[i.value for i in j] for j in load_ws.rows])
        header = data.iloc[0]
        data = data[1:]
        data.rename(columns=header, inplace=True)
        a_len = len(self.X)
        data.index = range(a_len, a_len + len(data))

        Rule = data[['source_name', 'condition_concept_id', 'concept_name']]

        comma = Rule['source_name'].str.replace(pat=',', repl='', regex=False)

        X = comma.str.replace(pat='-', repl=' ', regex=False)
        X2 = comma.str.replace(pat='-', repl='', regex=False)
        X = X.append(X2, ignore_index=False)
        X = X.str.lower()
        X.drop_duplicates(inplace=True)

        self.X = self.X.append(X, ignore_index=False)
        self.X.drop_duplicates(inplace=True)

        y = Rule[['condition_concept_id', 'concept_name']]
        self.y = self.y.append(y, ignore_index=False)
        self.y = self.y.loc[self.X.index]

        self.index_list = self.X.index.tolist()


        self.tfidf = CountVectorizer(stop_words=['***', ','], tokenizer=word_tokenize)

        self.tfidf_matrix = self.tfidf.fit_transform(self.X)

        self.should_not_use.extend(list(comma[data['should_not_use'] == 2].str.lower()))
        self.comment3.extend(list(comma[data['comment'] == 3].str.lower()))
        self.comment4.extend(list(comma[data['comment'] == 4].str.lower()))


    def OrderedSet(self, list):
        my_set = set()
        res = []
        for e in list:
            if e not in my_set:
                res.append(e)
                my_set.add(e)
        return res

    def Get_similar(self, statement=None):
        if statement == None: statement = list(self.X)
        concept_id = []
        concept_name = []
        #mapping_text = []
        for input_ in statement:
            input_ = re.sub(r'-', '', input_)
            input_ = self.tfidf.transform([input_.lower()])
            cosine_sim = cosine_similarity(self.tfidf_matrix, input_)
            sim_scores = list(enumerate(cosine_sim))
            sim_scores = sorted(sim_scores, key=lambda x: x[1], reverse=True)
            sim_scores = sim_scores[0]
            true_id = np.array(self.y[['condition_concept_id']].loc[self.index_list[sim_scores[0]]].dropna())
            true_name = np.array(self.y[['concept_name']].loc[self.index_list[sim_scores[0]]].dropna())
            #true_X = np.array(self.X.iloc[sim_scores[0]])
            concept_id.append(true_id)
            concept_name.append(true_name)
            #mapping_text.append(true_X)
        return concept_id, concept_name#, mapping_text

    def my_split(self, text, delimiter):
        token = []
        for i in range(len(text)):
            tmp = text[i].split(delimiter)
            for j in range(len(tmp)):
                token.append(tmp[j].strip().lower())
        return token

    def Check_if_any(self, statement=None):
        if statement == None: statement = list(self.X)
        concept_id = []
        concept_name = []
        #mapping_text = []

        for input_ in statement:
            input_ = re.sub(r'-', '', input_).lower()
            input_ = re.sub(r',', '', input_)
            out = False
            cont = False

            for tag2 in self.should_not_use:
                if tag2 in input_:
                    for tag3 in self.comment3:
                        if tag3 in input_:
                            cont = True
                    if cont:
                        break
                    concept_id.append([])
                    concept_name.append([])
                    #mapping_text.append(input_)
                    out = True
                    break
            if out:
                continue  # should not use 먼저 검사

            p = 0
            for all in self.X:
                if input_ == all:
                    allmatch_id = np.array(self.y[['condition_concept_id']].loc[self.index_list[p]].dropna())
                    allmatch_name = np.array(self.y[['concept_name']].loc[self.index_list[p]].dropna())
                    concept_id.append(allmatch_id)
                    concept_name.append(allmatch_name)
                    #mapping_text.append(all)
                    out = True
                    break
                p = p + 1
            if out:
                continue  # 전문 일치 검색

            ind = []
            p = 0
            # for not_tag2 in self.if_any_match:
            for not_tag4 in self.X:
                if not_tag4 in input_:
                    add = True
                    for com4 in self.comment4:
                        if com4 in input_:
                            add = False
                    if add:
                        ind.append(self.index_list[p])
                p = p + 1
            output1 = np.array(self.y[['condition_concept_id']].loc[ind].drop_duplicates())
            output2 = np.array(self.y[['concept_name']].loc[ind].drop_duplicates())
            n1 = []
            n2 = []
            for e in range(len(output1)):
                for l in output1[e]:
                    if l != None:
                        n1.append(int(l))
            for e in range(len(output2)):
                for l in output2[e]:
                    if l != None:
                        n2.append(l)

            concept_id.append(self.OrderedSet(n1))
            concept_name.append(self.OrderedSet(n2))
            #mapping_text.append(input_)
        return concept_id, concept_name#, mapping_text



if __name__ == "__main__" :
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    #myWindow.show()
    app.exec_()